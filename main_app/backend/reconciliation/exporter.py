# backend/reconciliation/exporter.py
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from io import BytesIO

def export_excel_bytes(df_total: pd.DataFrame, monthly_summary: pd.DataFrame) -> BytesIO:

    # --- Create workbook ---
    wb = Workbook()

    # 1️⃣ Reconciliation sheet (exact data as shown on screen)
    ws1 = wb.active
    ws1.title = "Reconciliation"

    # Write headers and rows directly
    ws1.append(list(df_total.columns))
    for row in df_total.itertuples(index=False):
        ws1.append(list(row))

    # Optional: Color code classification column if present
    if "Classification" in df_total.columns:
        fill_internal = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_incoming = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        fill_outgoing = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

        class_col = list(df_total.columns).index("Classification") + 1
        for row in range(2, ws1.max_row + 1):
            cls = ws1.cell(row=row, column=class_col).value
            if cls == "Internal":
                ws1.cell(row=row, column=class_col).fill = fill_internal
            elif cls == "Outgoing":
                ws1.cell(row=row, column=class_col).fill = fill_outgoing
            elif cls == "Incoming":
                ws1.cell(row=row, column=class_col).fill = fill_incoming

    # 2️⃣ Monthly Summary Sheet (already computed)
    if monthly_summary is not None and not monthly_summary.empty:
        ws2 = wb.create_sheet("Monthly Summary")
        ws2.append(list(monthly_summary.columns))
        for row in monthly_summary.itertuples(index=False):
            ws2.append(list(row))

        # Highlight “Grand Total” row if present
        for r in range(2, ws2.max_row + 1):
            if str(ws2.cell(r, 1).value).lower() == "grand total":
                for c in range(1, ws2.max_column + 1):
                    cell = ws2.cell(row=r, column=c)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")

    # --- Save to memory ---
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
