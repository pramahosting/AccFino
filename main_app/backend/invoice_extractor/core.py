"""
Invoice Extractor - core processing engine.
Adapted from test16.py: all extraction logic is preserved;
CLI / filesystem orchestration replaced with process_files().
"""

from __future__ import annotations

import io
import re
import shutil
import tempfile
import time
from pathlib import Path
from typing import List, Tuple

import numpy as np
import pandas as pd
import pdfplumber
from PIL import Image

try:
    import cv2
    _CV2_AVAILABLE = True
    _CV2_ERROR = None
except Exception as _e:
    _CV2_AVAILABLE = False
    _CV2_ERROR = str(_e)

try:
    from pdf2image import convert_from_path
    _PDF2IMAGE_AVAILABLE = True
    _PDF2IMAGE_ERROR = None
except Exception as _e:
    _PDF2IMAGE_AVAILABLE = False
    _PDF2IMAGE_ERROR = str(_e)

try:
    import pytesseract
    _TESSERACT_AVAILABLE = True
    _TESSERACT_ERROR = None
except Exception as _e:
    _TESSERACT_AVAILABLE = False
    _TESSERACT_ERROR = str(_e)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
    _OPENPYXL_ERROR = None
except Exception as _e:
    _OPENPYXL_AVAILABLE = False
    _OPENPYXL_ERROR = str(_e)

from datetime import datetime

IMG_EXTS = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp", ".webp"}
PDF_EXTS = {".pdf"}


# ============================================================
# -- SECTION 0: DATE / ABN / AMOUNT NORMALISERS --
# ============================================================

MONTH_MAP = {
    "jan": "01", "feb": "02", "mar": "03", "apr": "04",
    "may": "05", "jun": "06", "jul": "07", "aug": "08",
    "sep": "09", "oct": "10", "nov": "11", "dec": "12",
}


def normalise_date(raw, fallback_year=None) -> str:
    if not raw or str(raw).strip() in ("", "nan", "NaN"):
        return ""
    s = str(raw).strip()

    def validate(d, m, y) -> str:
        try:
            d, m, y = int(d), int(m), int(y)
        except (ValueError, TypeError):
            return ""
        if not (1 <= d <= 31 and 1 <= m <= 12 and 1900 <= y <= 2100):
            return ""
        return f"{d:02d}-{m:02d}-{y}"

    m = re.match(r"(\d{1,2})\s+([A-Za-z]+),?\s+(\d{4})", s)
    if m:
        mon = MONTH_MAP.get(m.group(2).lower()[:3], "")
        if mon:
            return validate(m.group(1), mon, m.group(3))

    m = re.match(r"([A-Za-z]+)\s+(\d{1,2}),?\s+(\d{4})", s)
    if m:
        mon = MONTH_MAP.get(m.group(1).lower()[:3], "")
        if mon:
            return validate(m.group(2), mon, m.group(3))

    m = re.match(r"(\d{4})[/\-\.](\d{1,2})[/\-\.](\d{1,2})$", s)
    if m:
        return validate(m.group(3), m.group(2), m.group(1))

    m = re.match(r"(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{4})$", s)
    if m:
        return validate(m.group(1), m.group(2), m.group(3))

    m = re.match(r"(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{2})$", s)
    if m:
        return validate(m.group(1), m.group(2), "20" + m.group(3))

    m = re.match(r"(\d{1,2})\s+([A-Za-z]{3})$", s)
    if m and fallback_year:
        mon = MONTH_MAP.get(m.group(2).lower()[:3], "")
        if mon:
            return validate(m.group(1), mon, str(fallback_year))

    return ""


def normalise_abn(raw) -> str:
    if not raw or str(raw).strip() in ("", "nan", "NaN"):
        return ""
    digits = re.sub(r"\D", "", str(raw))
    if len(digits) != 11:
        return ""
    weights = [10, 1, 3, 5, 7, 9, 11, 13, 15, 17, 19]
    total = sum((int(digits[i]) - (1 if i == 0 else 0)) * weights[i] for i in range(11))
    if total % 89 != 0:
        return ""
    return f"{digits[:2]} {digits[2:5]} {digits[5:8]} {digits[8:]}"


def normalise_amount(raw) -> str:
    if not raw or str(raw).strip() in ("", "nan", "NaN"):
        return ""
    s = str(raw).strip()
    s = re.sub(r"^(AU\$|AUD|\$)\s*", "", s, flags=re.IGNORECASE)
    if s.startswith("-"):
        return ""
    digits = re.sub(r"[^\d.]", "", s)
    if not digits or "." not in digits:
        return ""
    try:
        v = float(digits)
        if v <= 0 or v > 9_999_999:
            return ""
        return f"${v:,.2f}"
    except ValueError:
        return ""


def normalise_gst(raw) -> str:
    if not raw or str(raw).strip() in ("", "nan", "NaN"):
        return ""
    s = re.sub(r"^(AU\$|AUD|\$)\s*", "", str(raw).strip(), flags=re.IGNORECASE)
    if s.startswith("-"):
        return ""
    digits = re.sub(r"[^\d.]", "", s)
    if not digits or "." not in digits:
        return ""
    try:
        v = float(digits)
        if v <= 0 or v > 99_999:
            return ""
        return f"${v:,.2f}"
    except ValueError:
        return ""


# -- Per-column validators -----------------------------------------------------

_INV_BLACKLIST = {
    "DEBIT", "CREDIT", "PAYMENT", "BALANCE", "TOTAL", "CHANGE", "PURCHASE",
    "LIDCOME", "SINGAPORE", "INT", "DOM", "QANTAS", "EFT", "CARD", "GIFTCARD",
    "REGULAR", "AMERICAN", "XEXT", "TA", "CARI", "CARA", "X1", "CARDSURCHARGE",
}


def clean_invoice_number(raw) -> str:
    if not raw or str(raw).strip() in ("", "nan", "NaN"):
        return ""
    s = str(raw).strip().upper()
    if not re.search(r"\d", s):
        return ""
    if not re.match(r"^[A-Z0-9][A-Z0-9\-/\.]{1,29}$", s):
        return ""
    if s in _INV_BLACKLIST:
        return ""
    if re.match(r"^\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}$", s):
        return ""
    if re.match(r"^\d{11,}$", s):
        return ""
    return s


_ITEM_NO_BLACKLIST = {
    "PAYMENT", "BALANCE", "SINGAPORE", "INT", "QANTAS", "DOM", "GIFTCARD",
    "PURCHASE", "EFT", "CARD", "REGULAR", "CREDIT", "AMERICAN", "XEXT",
    "CARDSURCHARGE", "TA", "CARA", "X1", "HASALA", "RICE", "PICCOLO",
    "DEBIT", "LIDCOME",
}
_ITEM_NO_RE = re.compile(r"^[A-Z0-9][A-Z0-9\-/\.]{1,19}$", re.IGNORECASE)


def clean_item_number(raw) -> str:
    if not raw or str(raw).strip() in ("", "nan", "NaN"):
        return ""
    s = str(raw).strip().upper()
    if s in _ITEM_NO_BLACKLIST:
        return ""
    if re.match(r"^\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}$", s):
        return ""
    if not _ITEM_NO_RE.match(s):
        return ""
    if re.match(r"^[A-Z]{7,}$", s):
        return ""
    return s


def clean_qty(raw) -> str:
    if raw is None or str(raw).strip() in ("", "nan", "NaN"):
        return ""
    try:
        v = float(str(raw).strip().replace(",", ""))
        if v <= 0 or v > 100_000:
            return ""
        return str(int(v)) if v == int(v) else str(round(v, 4))
    except ValueError:
        return ""


_OCR_JUNK_RE = re.compile(
    r"^[^a-zA-Z0-9]*$"
    r"|^[oe\s_\-\.\|=\[\]\"'~`^]{1,}$"
    r"|^\W{2,}",
    re.IGNORECASE,
)


def clean_company(raw) -> str:
    if not raw or str(raw).strip() in ("", "nan", "NaN"):
        return ""
    s = str(raw).strip()
    if len(s) < 3:
        return ""
    if _OCR_JUNK_RE.match(s):
        return ""
    s_clean = re.sub(r"^[\W_]+", "", s).strip()
    if not s_clean:
        return ""
    alnum = sum(1 for c in s_clean if c.isalnum())
    if alnum / max(len(s_clean), 1) < 0.40:
        return ""
    words = s_clean.split()
    if all(len(w) <= 3 for w in words) and len(words) <= 4:
        return ""
    if len(words) == 1 and len(s_clean) <= 4 and not re.match(r"^[A-Z]{2,4}$", s_clean):
        return ""
    if re.search(r"Depart\s+Date|Depart\s+Time", s, re.IGNORECASE):
        return ""
    if re.match(r"Property\s+GST\s+ID", s, re.IGNORECASE):
        return ""
    unique_chars = set(s_clean.lower().replace(" ", ""))
    if len(unique_chars) <= 3 and len(s_clean) > 4:
        return ""
    if re.match(r"^\d+\s+\w+\s+(?:st|rd|ave|dr|ln|blvd|street|road|avenue|drive|lane|way|place|court|crescent)\b",
                s_clean, re.IGNORECASE):
        return ""
    return s_clean.rstrip(",:;.!?-")


_DESC_JUNK_RE = re.compile(
    r"^\s*(?:\*+\s*total\s+(?:excluding|including)?\s*gst"
    r"|\*+\s*total\s+gst"
    r"|total\s*:"
    r"|change\s*:"
    r"|taks?\s*:"
    r"|total\s+price"
    r"|charged\s+to\s+[ac]x\*"
    r"|by\s+[ac]x\*"
    r"|by\s+[ac]a\*"
    r"|\d+[\.,]\d{2}\s+\d+[\.,]\d{2})"
    r"|^\s*[\$\|\-\=\+\_\~\^\`\'\"\[\]\{\}]{2,}"
    r"|^\s*[a-z0-9]{1,2}\s*$",
    re.IGNORECASE,
)
_DROP_DESC_RE = re.compile(
    r"^\s*(?:\*+\s*total|total\s+gst|total\s+excluding|total\s+including"
    r"|payment\s+received|amount\s+paid|balance\s+due|change\s+given"
    r"|charged\s+to|by\s+[ac][xa]\*|rrn\b|appr?ov|authoris)"
    r"|^\s*[\d\s\$\.,]+$",
    re.IGNORECASE,
)


def clean_item_description(raw) -> str:
    if not raw or str(raw).strip() in ("", "nan", "NaN"):
        return ""
    s = str(raw).strip()
    if _DESC_JUNK_RE.match(s):
        return ""
    s = re.sub(r"[|\[\]~`^\\]", "", s).strip()
    s = re.sub(r"\s{2,}", " ", s)
    if len(s) < 2:
        return ""
    if not re.search(r"[a-zA-Z]", s):
        return ""
    return s


def validate_unit_price_consistency(unit_price: str, qty: str, amount: str) -> str:
    if not unit_price:
        return ""

    def to_f(s):
        try:
            return float(re.sub(r"[^\d.]", "", str(s)))
        except Exception:
            return None

    up = to_f(unit_price)
    if up is None or up <= 0:
        return ""
    amt = to_f(amount)
    q = to_f(qty)
    if amt is not None and amt > 0:
        if q is not None and q > 0:
            calc = up * q
            if abs(calc - amt) / amt > 0.015:
                return ""
        else:
            if up > amt * 1.01:
                return ""
    return unit_price


def validate_gst_vs_total(gst: str, invoice_total: str) -> str:
    if not gst:
        return ""

    def to_f(s):
        try:
            return float(re.sub(r"[^\d.]", "", str(s)))
        except Exception:
            return None

    g = to_f(gst)
    t = to_f(invoice_total)
    if g is None or g <= 0:
        return ""
    if t is not None and t > 0:
        if g > t:
            return ""
        expected = t / 11.0
        if expected > 0 and (g / expected) > 1.5:
            return ""
    return gst


def _apply_column_validators(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Invoice number"] = df["Invoice number"].apply(clean_invoice_number)
    df["Invoice date"] = df["Invoice date"].apply(normalise_date)
    df["ABN"] = df["ABN"].apply(normalise_abn)
    df["Qty"] = df["Qty"].apply(clean_qty)
    df["Company"] = df["Company"].apply(clean_company)
    df["Item description"] = df["Item description"].apply(clean_item_description)
    df["Amount"] = df["Amount"].apply(normalise_amount)
    df["GST"] = df["GST"].apply(normalise_gst)
    df["Invoice total"] = df["Invoice total"].apply(normalise_amount)
    df["Unit Price"] = df["Unit Price"].apply(normalise_amount)

    df["Unit Price"] = df.apply(
        lambda r: validate_unit_price_consistency(r["Unit Price"], r["Qty"], r["Amount"]),
        axis=1,
    )
    df["GST"] = df.apply(
        lambda r: validate_gst_vs_total(r["GST"], r["Invoice total"]),
        axis=1,
    )

    empty_desc = df["Item description"].str.strip() == ""
    empty_amt = df["Amount"].str.strip() == ""
    df = df[~(empty_desc & empty_amt)].reset_index(drop=True)

    junk_desc = df["Item description"].apply(
        lambda x: bool(_DROP_DESC_RE.match(str(x))) if str(x).strip() else False
    )
    df = df[~junk_desc].reset_index(drop=True)
    df = df.fillna("")
    return df


# ============================================================
# -- SECTION A: BANK STATEMENT LOGIC --
# ============================================================

BANK_SIGNATURES = {
    "ing":       [r"\bing\b", r"ing direct", r"orange everyday", r"ing bank",
                  r"mortgage simplifier", r"living super.*ing"],
    "macquarie": [r"macquarie bank", r"macquarie\.com\.au", r"\bmacquarie\b"],
    "bankwest":  [r"bankwest", r"bankwest\.com\.au"],
    "suncorp":   [r"suncorp", r"suncorp\.com\.au"],
    "stgeorge":  [r"st\.george", r"stgeorge\.com\.au"],
    "bendigo":   [r"bendigo bank", r"bendigobank\.com\.au"],
    "amp":       [r"\bamp bank\b", r"amp\.com\.au", r"\bamp\b.*bett3r", r"bett3r account"],
    "bom":       [r"bank of melbourne", r"bankofmelbourne\.com\.au"],
    "mystate":   [r"mystate bank", r"mystate\.com\.au", r"\bmystate\b"],
    "newcastle": [r"newcastle permanent", r"newcastlepermanent\.com\.au"],
    "pnbank":    [r"p&n bank", r"pnbank\.com\.au"],
    "hsbc":      [r"\bhsbc\b", r"hsbc bank australia", r"hsbc\.com\.au", r"hsbc.*home value"],
    "nab":       [r"\bnab\b", r"national australia bank", r"nab\.com\.au"],
    "westpac":   [r"westpac", r"rocket statement", r"westpac\.com\.au",
                  r"altitude black", r"altitude.*mastercard"],
    "commbank":  [r"commonwealth bank", r"commbank", r"netbank", r"cba\.com\.au",
                  r"commonwealthbank"],
    "anz":       [r"australia and new zealand banking", r"\banz\b", r"anz\.com\.au",
                  r"anz.*business one zero", r"anz.*v2 plus", r"anz.*access advantage"],
}

ACCOUNT_TYPE_SIGNATURES = {
    "cheque":         [r"cheque account", r"business cheque"],
    "offset":         [r"everyday offset", r"offset account"],
    "superannuation": [r"living super", r"superannuation", r"accumulation\s+phase"],
    "term_deposit":   [r"term investment", r"maturity statement"],
    "savings":        [r"savings account", r"netbank saver", r"online saver",
                       r"savings maximiser", r"esaver", r"reward saver",
                       r"bump savings", r"bonus saver", r"cash management account",
                       r"bett3r"],
    "transaction":    [r"transaction account", r"everyday account", r"smart access",
                       r"orange everyday", r"everyday global", r"complete freedom",
                       r"anz plus", r"access advantage", r"classic banking",
                       r"visa debit", r"choice account", r"nab classic",
                       r"business one zero", r"v2 plus", r"hatch spending",
                       r"business\s+trans\s+act", r"business\s+transaction"],
    "home_loan":      [r"home loan account", r"rocket repay", r"mortgage simplifier",
                       r"home value loan", r"variable rate home loan",
                       r"fixed rate home loan", r"personal loan",
                       r"standard variable home loan"],
    "credit_card":    [r"credit card", r"mastercard", r"visa card",
                       r"rewards.*card", r"altitude.*mastercard",
                       r"amplify.*credit", r"low rate.*visa"],
}


def classify_bank(text: str) -> tuple:
    t = text.lower()
    header_text = "\n".join(t.split("\n")[:25])
    bank = "unknown"
    for b, patterns in BANK_SIGNATURES.items():
        if any(re.search(p, header_text) for p in patterns):
            bank = b
            break
    if bank == "unknown":
        for b, patterns in BANK_SIGNATURES.items():
            if any(re.search(p, t) for p in patterns):
                bank = b
                break
    acct = "unknown"
    for a, patterns in ACCOUNT_TYPE_SIGNATURES.items():
        if any(re.search(p, t) for p in patterns):
            acct = a
            break
    return bank, acct


_BANK_CONFIRM = [
    re.compile(r"\bstatement\s+period\b", re.IGNORECASE),
    re.compile(r"\bopening\s+balance\b", re.IGNORECASE),
    re.compile(r"\bclosing\s+balance\b", re.IGNORECASE),
    re.compile(r"\baccount\s+(?:number|no\.?)\b.*?\d{4}", re.IGNORECASE | re.DOTALL),
    re.compile(r"\bbsb\b\s*:?\s*\d{3}", re.IGNORECASE),
    re.compile(r"\btransaction\s+(?:date|details|history)\b", re.IGNORECASE),
    re.compile(
        r"\b(?:debit|withdrawal)\b.{0,40}\b(?:credit|deposit)\b.{0,40}\bbalance\b",
        re.IGNORECASE | re.DOTALL,
    ),
    re.compile(r"\bavailable\s+(?:balance|funds)\b", re.IGNORECASE),
    re.compile(r"\bminimum\s+(?:monthly\s+)?payment\b", re.IGNORECASE),
    re.compile(r"\bstatement\s+(?:number|no\.?|date)\b", re.IGNORECASE),
    re.compile(r"\bbank\s+statement\b", re.IGNORECASE),
    re.compile(r"\bdirect\s+debit\b", re.IGNORECASE),
]

_BANK_DISQUALIFY = [
    re.compile(r"\bterminal\s+id\b", re.IGNORECASE),
    re.compile(r"\bmerchant\s+id\b", re.IGNORECASE),
    re.compile(r"\bacquirer\b", re.IGNORECASE),
    re.compile(r"\bauth(?:orisation)?\s+id\b", re.IGNORECASE),
    re.compile(r"#{4,}\d{4}"),
    re.compile(r"\brrn\b\s*:?\s*\d{6,}", re.IGNORECASE),
    re.compile(r"\bapproved\s+0{2}\b", re.IGNORECASE),
    re.compile(r"\barqc\b", re.IGNORECASE),
    re.compile(r"^\s*nab\s+eftpos\s*$", re.IGNORECASE | re.MULTILINE),
    re.compile(r"\btax\s+invoice\b", re.IGNORECASE),
    re.compile(r"\breceipt\s+no\b", re.IGNORECASE),
    re.compile(r"\bthank\s+you\s+for\s+(?:shopping|your\s+purchase)\b", re.IGNORECASE),
    re.compile(r"\bgst\s+included\b", re.IGNORECASE),
    re.compile(r"\bcustomer\s+copy\b", re.IGNORECASE),
    re.compile(r"\bchange\b.{0,15}\$\s*0\.00", re.IGNORECASE),
    re.compile(r"\bnumber\s+of\s+items\b", re.IGNORECASE),
]


def is_bank_statement(text: str) -> bool:
    bank, _ = classify_bank(text)
    if bank == "unknown":
        return False
    if any(p.search(text) for p in _BANK_DISQUALIFY):
        return False
    confirm_hits = sum(1 for p in _BANK_CONFIRM if p.search(text))
    return confirm_hits >= 2


_AU_STATES_PAT = r"(?:NSW|VIC|QLD|SA|WA|TAS|NT|ACT)"
_NON_ADDR_RE = re.compile(
    r"bank|statement|account|bsb|abn|afsl|phone|fax|email|www\.|\\.com|"
    r"branch|interest|balance|transaction|period|date|debit|credit|"
    r"subject to|authorised|disclaimer|conditions|privacy",
    re.IGNORECASE,
)
_BANK_HEADER_RE = re.compile(
    r"\b(commonwealth|westpac|national australia|mystate|bendigo|suncorp|"
    r"bankwest|macquarie|ing direct|hsbc|amplify|altitude|netbank|"
    r"access advantage|smart access|business one zero|v2 plus|"
    r"reward saver|incentive saver|bonus saver|life account|bett3r|"
    r"mortgage simplifier|home value|living super|cash management|"
    r"everyday account|business everyday|hatch spending|"
    r"term investment|maturity statement)\b",
    re.IGNORECASE,
)
_TITLE_RE = re.compile(r"^(Mr\.?|Mrs\.?|Ms\.?|Miss\.?|Dr\.?|Prof\.?)\s+", re.IGNORECASE)
_ALLCAPS_TITLE_RE = re.compile(r"^(MR\.?|MRS\.?|MS\.?|MISS\.?|DR\.?)\s+[A-Z]")


def _is_name_line(line):
    if _NON_ADDR_RE.search(line):
        return False
    if _BANK_HEADER_RE.search(line):
        return False
    if _TITLE_RE.match(line) or _ALLCAPS_TITLE_RE.match(line):
        return True
    if (re.match(r"^[A-Z][A-Z\s&]+$", line) and 2 <= len(line.split()) <= 6
            and not re.search(r"\d", line)
            and not re.search(r"\b(BANK|STATEMENT|ACCOUNT|BSB|ABN|PERIOD|DATE)\b", line)):
        return True
    if (re.match(r"^([A-Z][a-z]+\s+){1,3}[A-Z][a-z]+$", line)
            and not _BANK_HEADER_RE.search(line) and not re.search(r"\d", line)):
        return True
    return False


def _is_address_line(line):
    if _NON_ADDR_RE.search(line):
        return False
    if re.match(r"^\d+[\w/]*\s+\w", line):
        return True
    if re.match(r"^(Unit|Level|Shop|Suite|Floor|Flat|PO Box|GPO Box|Locked Bag)\b", line, re.IGNORECASE):
        return True
    if re.search(r"\b" + _AU_STATES_PAT + r"\s+\d{4}\b", line, re.IGNORECASE):
        return True
    return False


def _extract_holder_and_address(text):
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    search_lines = lines[:70]
    holder, address_parts = "", []
    dear_re = re.compile(
        r"[Dd]ear\s+(?:(?:Mr\.?|Mrs\.?|Ms\.?|Miss\.?|Dr\.?)\s+)?"
        r"([A-Z][a-zA-Z]+(?:\s+[A-Z][a-zA-Z]+){0,3}),?",
        re.IGNORECASE,
    )
    for line in search_lines:
        m = dear_re.search(line)
        if m:
            candidate = m.group(1).strip()
            if not _BANK_HEADER_RE.search(candidate):
                holder = candidate.title()
                break
    name_line_idx = None
    for i, line in enumerate(search_lines):
        if _is_name_line(line):
            candidate = line.title()
            if not holder:
                holder = candidate
            name_line_idx = i
            break
    if name_line_idx is not None:
        for line in search_lines[name_line_idx + 1: name_line_idx + 6]:
            if _is_address_line(line):
                address_parts.append(line)
            elif address_parts:
                break
    if not address_parts:
        postcode_re = re.compile(r"\b" + _AU_STATES_PAT + r"\s+\d{4}\b", re.IGNORECASE)
        for i, line in enumerate(search_lines):
            if postcode_re.search(line):
                start = max(0, i - 3)
                if name_line_idx is not None:
                    start = max(name_line_idx + 1, start)
                for addr_line in search_lines[start: i + 1]:
                    if _is_address_line(addr_line) or postcode_re.search(addr_line):
                        address_parts.append(addr_line)
                break
    seen = []
    for part in address_parts:
        if part not in seen:
            seen.append(part)
    address = ", ".join(seen)
    if not holder:
        for line in search_lines[:25]:
            words = line.split()
            if (2 <= len(words) <= 5 and all(w[0].isupper() for w in words if w)
                    and not re.search(r"\d", line)
                    and not _NON_ADDR_RE.search(line)
                    and not _BANK_HEADER_RE.search(line)):
                holder = line
                break
    return holder.strip(), address.strip()


def extract_meta(text, bank, acct):
    t = text.lower()
    acct_num = ""
    for pattern in [
        r"account\s+(?:number|no\.?)[:\s]+([0-9\s\-x*]{6,25})",
        r"account:\s*([a-z0-9\s\-x*]{6,30})",
        r"([0-9]{3,4}[\s\-][0-9]{3,6}[\s\-][0-9]{6,10})",
        r"([0-9]{4}\s[0-9]{4}\s[0-9]{4}\s[0-9]{4})",
    ]:
        m = re.search(pattern, t)
        if m:
            acct_num = m.group(1).strip()
            break
    period_start, period_end = "", ""
    for pattern in [
        r"(-:statement\s+)-period[:\s]+(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\s*[--to]+\s*(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})",
        r"(-:statement\s+)-period[:\s]+(\d{1,2}\s+\w{3}\s+\d{4})\s+[--to]+\s*(\d{1,2}\s+\w{3}\s+\d{4})",
        r"(\d{1,2}\s+\w{3}\s+\d{4})\s*(-:to|-|-)\s*(\d{1,2}\s+\w{3}\s+\d{4})",
        r"(\d{1,2}/\d{1,2}/\d{4})\s*[--]\s*(\d{1,2}/\d{1,2}/\d{4})",
        r"(\d{1,2}\s+[a-z]{3}\s+\d{2,4})\s+to\s+(\d{1,2}\s+[a-z]{3}\s+\d{2,4})",
    ]:
        m = re.search(pattern, t)
        if m:
            period_start = m.group(1).strip()
            period_end = m.group(2).strip()
            break
    if not period_start:
        all_dates = re.findall(r"\b(\d{2}/\d{2}/\d{4})\b", text)
        if len(all_dates) >= 2:
            period_start = all_dates[-1]
            period_end = all_dates[0]
    period_start = normalise_date(period_start)
    period_end = normalise_date(period_end)
    holder, address = _extract_holder_and_address(text)
    return {
        "bank": bank, "account_type": acct, "account_number": acct_num,
        "period_start": period_start, "period_end": period_end,
        "account_holder": holder, "address": address,
    }


def parse_amount(raw):
    if raw is None:
        return None, None
    s = str(raw).strip()
    if not s or s in ("-", "-"):
        return None, None
    direction = None
    if s.endswith("-"):
        direction, s = "debit", s[:-1]
    elif s.endswith("+"):
        direction, s = "credit", s[:-1]
    if direction is None:
        if s.startswith("-"):
            direction, s = "credit", s[1:]
        elif s.startswith("+"):
            direction, s = "credit", s[1:]
    s = re.sub(r"[$,\s]", "", s)
    try:
        return float(s), direction
    except ValueError:
        return None, None


def is_date(s):
    if s is None:
        return False
    s = str(s).strip()
    return bool(
        re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}$", s)
        or re.match(r"^\d{1,2}/\d{1,2}$", s)
        or re.match(r"^\d{1,2}\s+[A-Za-z]{3}\s+\d{4}$", s)
        or re.match(r"^\d{1,2}\s+[A-Za-z]{3}$", s)
    )


SKIP_RE = re.compile(
    r"^(opening balance|closing balance|statement opening|statement closing|"
    r"brought forward|carried forward|total\b|subtotal|minimum payment|"
    r"payment due|credit limit|available funds|date\s+desc|transaction\s+det|"
    r"processed\s+trans|card\s+used|transaction details)",
    re.IGNORECASE,
)


def should_skip(text):
    return bool(SKIP_RE.match(str(text).strip()))


def clean_desc(raw):
    if raw is None:
        return ""
    return " ".join(str(raw).split())


def clean_balance(raw):
    if not raw:
        return None
    s = re.sub(r"\s*(CR|DR|C|D)\s*$", "", str(raw).strip(), flags=re.IGNORECASE)
    s = re.sub(r"[$,\s]", "", s)
    try:
        return float(s)
    except ValueError:
        return None


HEADER_MAP = {
    "date": "date", "processed": "date", "transaction date": "date",
    "txn date": "date", "value date": "date",
    "description": "description", "transaction details": "description",
    "transaction detail": "description", "details": "description",
    "narration": "description", "particulars": "description",
    "narrative": "description", "description of transaction": "description",
    "transaction": "description", "transaction description": "description",
    "card used": "card", "card": "card",
    "debit": "debit", "debits": "debit", "withdrawal": "debit",
    "withdrawals": "debit", "money out": "debit", "dr": "debit",
    "debit ($)": "debit", "withdrawals ($)": "debit",
    "charges": "debit", "charges ($)": "debit", "fees": "debit",
    "interest charged": "debit", "amount charged": "debit",
    "interest/fees": "debit", "interest/fees ($)": "debit",
    "credit": "credit", "credits": "credit", "deposit": "credit",
    "deposits": "credit", "money in": "credit", "cr": "credit",
    "credit ($)": "credit", "deposits ($)": "credit",
    "repayment": "credit", "repayments": "credit",
    "payment made": "credit", "amount paid": "credit",
    "amount": "amount", "amount (aud)": "amount", "transactions": "amount",
    "balance": "balance", "running balance": "balance",
    "balance ($)": "balance", "loan balance": "balance",
    "account balance": "balance",
    "reference": "reference", "ref": "reference",
    "cheque no": "reference", "receipt no": "reference",
}


def resolve_headers(raw_headers):
    result = {}
    for i, h in enumerate(raw_headers):
        if h is None:
            continue
        key = " ".join(str(h).split()).lower()
        key = re.sub(r"\s*\([a-z]{3}\)\s*$", "", key).strip()
        key = re.sub(r"\([^)]*\)$", "", key).strip()
        concept = HEADER_MAP.get(key)
        if concept and concept not in result:
            result[concept] = i
    return result


def get_cell(row, idx, default=""):
    if idx is None or idx >= len(row):
        return default
    v = row[idx]
    return clean_desc(v) if v is not None else default


def parse_table(table, bank, acct, fallback_year=None):
    if not table or len(table) < 2:
        return []
    header_row_idx, col_map = 0, {}
    for ri, row in enumerate(table[:8]):
        if row is None:
            continue
        normalised = [" ".join(str(c).split()).lower() if c else "" for c in row]
        cm = resolve_headers(normalised)
        if len(cm) >= 2:
            col_map, header_row_idx = cm, ri
            break
    if not col_map:
        return []
    has_sep = "debit" in col_map and "credit" in col_map
    has_amount = "amount" in col_map
    transactions = []
    current_year = fallback_year
    for row in table[header_row_idx + 1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        first_cell = str(row[0]).strip() if row[0] else ""
        if re.fullmatch(r"(19|20)\d{2}", first_cell):
            current_year = int(first_cell)
            continue
        date_raw = get_cell(row, col_map.get("date"))
        if not date_raw or not is_date(date_raw):
            if is_date(first_cell):
                date_raw = first_cell
            else:
                continue
        desc = get_cell(row, col_map.get("description"))
        if should_skip(desc) or should_skip(date_raw):
            continue
        if not desc:
            for c in row[1:]:
                if c and str(c).strip():
                    desc = clean_desc(c)
                    break
        bal_val = clean_balance(get_cell(row, col_map.get("balance")))
        ref = get_cell(row, col_map.get("reference"))
        if has_sep:
            deb_val, _ = parse_amount(get_cell(row, col_map.get("debit")))
            cre_val, _ = parse_amount(get_cell(row, col_map.get("credit")))
            debit = abs(deb_val) if deb_val else None
            credit = abs(cre_val) if cre_val else None
        elif has_amount:
            raw_amt = get_cell(row, col_map.get("amount"))
            val, direction = parse_amount(raw_amt)
            if val is None:
                continue
            s_amt = str(raw_amt).strip()
            leading_neg = s_amt.startswith("-")
            leading_pos = s_amt.startswith("+")
            if direction == "debit":
                debit, credit = abs(val), None
            elif direction == "credit" and not leading_neg and not leading_pos:
                debit, credit = None, abs(val)
            elif leading_pos:
                debit, credit = None, abs(val)
            elif leading_neg:
                if acct == "credit_card":
                    debit, credit = None, abs(val)
                else:
                    debit, credit = abs(val), None
            elif acct == "credit_card":
                debit, credit = abs(val), None
            else:
                debit, credit = None, abs(val)
        else:
            continue
        if debit is None and credit is None:
            continue
        transactions.append({
            "date": normalise_date(date_raw, fallback_year=current_year),
            "description": desc, "reference": ref,
            "debit": debit, "credit": credit, "balance": bal_val,
        })
    return transactions


def _extract_year_context(text):
    for line in text.splitlines():
        line = line.strip()
        m = re.fullmatch(r"(19|20)\d{2}", line)
        if m:
            return int(m.group(0))
    m = re.search(r"\b((?:19|20)\d{2})\b", text)
    return int(m.group(1)) if m else None


def parse_anz_credit_text(text, fallback_year=None):
    transactions = []
    line_re = re.compile(
        r"(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})\s+(\d{4})\s+(.+?)\s+(\$[\d,]+\.\d{2}[+\-])"
    )
    for line in text.splitlines():
        m = line_re.search(line.strip())
        if not m:
            continue
        desc = clean_desc(m.group(4))
        if should_skip(desc):
            continue
        val, direction = parse_amount(m.group(5))
        if val is None:
            continue
        debit = abs(val) if direction == "debit" else None
        credit = abs(val) if direction == "credit" else None
        if debit is None and credit is None:
            debit = abs(val)
        transactions.append({
            "date": normalise_date(m.group(2), fallback_year=fallback_year),
            "description": desc, "reference": "",
            "debit": debit, "credit": credit, "balance": None,
        })
    return transactions


def parse_westpac_rocket_text(text, fallback_year=None):
    transactions = []
    raw_lines = [l.strip() for l in text.splitlines()]
    pending_date, pending_desc, pending_year = None, "", ""
    _ctx_year = fallback_year

    def flush(extra=""):
        nonlocal pending_date, pending_desc
        if not pending_date:
            return
        combined = f"{pending_desc} {extra}".strip()
        if not combined or should_skip(combined):
            pending_date = pending_desc = ""
            return
        nums = re.findall(r"-?[\d,]+\.\d{2}", combined)
        desc = re.sub(r"-?[\d,]+\.\d{2}", "", combined)
        desc = re.sub(r"\s{2,}", " ", desc).strip()
        if len(nums) >= 2:
            amt = float(nums[-2].replace(",", ""))
            bal = float(nums[-1].replace(",", ""))
            is_wdl = bool(re.search(r"withdrawal|payment|eftpos|atm|authority to", desc, re.IGNORECASE))
            debit, credit = (abs(amt), None) if (is_wdl or amt < 0) else (None, abs(amt))
            if desc and not should_skip(desc):
                _yr = int(pending_year) if pending_year else _ctx_year
                transactions.append({
                    "date": normalise_date(pending_date, fallback_year=_yr),
                    "description": desc, "reference": "",
                    "debit": debit, "credit": credit, "balance": bal,
                })
        pending_date = pending_desc = ""

    for line in raw_lines:
        if not line:
            continue
        if re.match(r"^\d{4}$", line):
            pending_year = line
            continue
        dm = re.match(r"^(\d{1,2}\s+[A-Za-z]{3})\s*(.*)", line)
        if dm:
            flush()
            pending_date = f"{dm.group(1)} {pending_year}".strip() if pending_year else dm.group(1)
            pending_desc = dm.group(2).strip()
            if re.search(r"-?[\d,]+\.\d{2}\s+-?[\d,]+\.\d{2}\s*$", pending_desc):
                flush()
            continue
        if pending_date:
            pending_desc = f"{pending_desc} {line}".strip()
            if re.search(r"-?[\d,]+\.\d{2}\s+-?[\d,]+\.\d{2}\s*$", pending_desc):
                flush()
    flush()
    return transactions


def _try_cba_amount_patterns(rest):
    pat_a = re.search(r"([\d,]+\.\d{2})\s+\$\s+\$([\d,]+\.\d{2})\s*(?:CR)?\s*$", rest, re.I)
    pat_b = re.search(r"\$([\d,]+\.\d{2})\s+\$([\d,]+\.\d{2})\s*(?:CR)?\s*$", rest, re.I)
    pat_c = re.search(r"^([\d,]+\.\d{2})\s+\$([\d,]+\.\d{2})\s*(?:CR)?\s*$", rest, re.I)
    for pat, is_debit in [(pat_a, True), (pat_b, False), (pat_c, True)]:
        if pat:
            return float(pat.group(1).replace(",", "")), float(pat.group(2).replace(",", "")), is_debit
    return None


def parse_cba_text(text, fallback_year=None):
    transactions = []
    lines = [l.strip() for l in text.splitlines()]
    i, current_date, current_desc = 0, None, None
    _ctx_year = _extract_year_context(text) or fallback_year

    def emit(desc, rest_line, date):
        if should_skip(desc):
            return False
        hit = _try_cba_amount_patterns(rest_line)
        if hit:
            amt, bal, is_debit = hit
            rest_label = re.sub(r"[\d,]+\.\d{2}", "", rest_line)
            rest_label = re.sub(r"[$CR\s]+", " ", rest_label, flags=re.IGNORECASE).strip()
            full_desc = f"{desc} {rest_label}".strip() if rest_label else desc
            transactions.append({
                "date": normalise_date(date, fallback_year=_ctx_year),
                "description": full_desc, "reference": "",
                "debit": amt if is_debit else None,
                "credit": None if is_debit else amt, "balance": bal,
            })
            return True
        return False

    while i < len(lines):
        line = lines[i]
        if not line:
            i += 1
            continue
        dm = re.match(r"^(\d{1,2}\s+[A-Za-z]{3}(?:\s+\d{4})?)\s+(.*)", line)
        if dm:
            current_date = dm.group(1).strip()
            rest_of_line = dm.group(2).strip()
            if emit(rest_of_line, rest_of_line, current_date):
                current_date = current_desc = None
            else:
                current_desc = rest_of_line
            i += 1
            continue
        if current_date and current_desc is not None:
            if should_skip(current_desc):
                current_date = current_desc = None
                i += 1
                continue
            if emit(current_desc, line, current_date):
                current_date = current_desc = None
            else:
                current_desc = f"{current_desc} {line}".strip()
        i += 1
    return transactions


def parse_hsbc_credit_text(text, fallback_year=None):
    transactions = []
    line_re = re.compile(r"^(\d{1,2}/\d{2}/\d{4})\s+(?:\d{4}\s+)?(.+?)\s+(-?\$[\d,]+\.\d{2})\s*$")
    for line in text.splitlines():
        m = line_re.match(line.strip())
        if not m:
            continue
        desc = clean_desc(m.group(2))
        amt_raw = m.group(3)
        if should_skip(desc):
            continue
        is_neg = amt_raw.startswith("-")
        try:
            val = float(re.sub(r"[$,]", "", amt_raw.lstrip("-")))
        except ValueError:
            continue
        debit, credit = (None, val) if is_neg else (val, None)
        transactions.append({
            "date": normalise_date(m.group(1), fallback_year=fallback_year),
            "description": desc, "reference": "",
            "debit": debit, "credit": credit, "balance": None,
        })
    return transactions


def parse_anz_cheque_text(text, fallback_year=None):
    transactions = []
    prev_balance = None
    line_re = re.compile(
        r"^(\d{2}/\d{2}/\d{4})\s+(.+?)\s*\$([\d,]+\.\d{2})\s*\$([\d,]+\.\d{2})\s*$"
    )
    credit_kw = re.compile(
        r"transfer from|salary|dividend|deposit|credit|received|refund|income|rental",
        re.IGNORECASE,
    )
    for line in text.splitlines():
        m = line_re.match(line.strip())
        if not m:
            continue
        date_raw = m.group(1)
        desc = clean_desc(m.group(2))
        amt = float(m.group(3).replace(",", ""))
        bal = float(m.group(4).replace(",", ""))
        if should_skip(desc):
            continue
        if prev_balance is not None:
            delta = bal - prev_balance
            if abs(delta + amt) < 0.02:
                debit, credit = amt, None
            elif abs(delta - amt) < 0.02:
                debit, credit = None, amt
            else:
                debit, credit = (None, amt) if credit_kw.search(desc) else (amt, None)
        else:
            debit, credit = (None, amt) if credit_kw.search(desc) else (amt, None)
        prev_balance = bal
        transactions.append({
            "date": normalise_date(date_raw, fallback_year=fallback_year),
            "description": desc, "reference": "",
            "debit": debit, "credit": credit, "balance": bal,
        })
    return transactions


def parse_generic_text(text, acct):
    transactions = []
    two_amt_re = re.compile(r"(-?\$?[\d,]+\.\d{2}[+\-]?)\s+(-?\$?[\d,]+\.\d{2}[+\-]?)\s*$")
    one_amt_re = re.compile(r"(-?\$?[\d,]+\.\d{2}[+\-]?)\s*$")
    credit_kw = re.compile(
        r"salary|direct credit|interest credit|transfer from|deposit|refund|credit",
        re.IGNORECASE,
    )
    fallback_year = _extract_year_context(text)
    current_year = fallback_year
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        if re.fullmatch(r"(19|20)\d{2}", line):
            current_year = int(line)
            continue
        dm = re.match(
            r"^(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}|\d{1,2}\s+[A-Za-z]{3}(?:\s+\d{4})?)\s+(.*)",
            line,
        )
        if not dm:
            continue
        date_raw, rest = dm.group(1), dm.group(2)
        if should_skip(rest):
            continue
        tm = two_amt_re.search(rest)
        if tm:
            raw_amt = tm.group(1)
            raw_bal = tm.group(2)
            desc = rest[: tm.start()].strip()
            if not desc:
                continue
            bal_val = clean_balance(re.sub(r"[\$,]", "", raw_bal))
        else:
            am = one_amt_re.search(rest)
            if not am:
                continue
            raw_amt = am.group(1)
            bal_val = None
            desc = rest[: am.start()].strip()
            if not desc:
                continue
        if should_skip(desc):
            continue
        leading_neg = str(raw_amt).strip().startswith("-")
        val, direction = parse_amount(raw_amt)
        if val is None:
            continue
        if direction == "debit":
            debit, credit = abs(val), None
        elif direction == "credit":
            debit, credit = None, abs(val)
        elif leading_neg and acct == "credit_card":
            debit, credit = None, abs(val)
        elif acct == "credit_card":
            debit, credit = abs(val), None
        else:
            if credit_kw.search(desc):
                debit, credit = None, abs(val)
            else:
                debit, credit = abs(val), None
        transactions.append({
            "date": normalise_date(date_raw, fallback_year=current_year),
            "description": clean_desc(desc),
            "reference": "", "debit": debit, "credit": credit, "balance": bal_val,
        })
    return transactions


def extract_from_pdf_bank(pdf_path):
    all_text, page_tables = "", []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            all_text += (page.extract_text() or "") + "\n"
            page_tables.extend(page.extract_tables() or [])
    bank, acct = classify_bank(all_text)
    meta = extract_meta(all_text, bank, acct)
    yr = _extract_year_context(all_text)
    single_row_data = [
        t for t in page_tables
        if t and len(t) == 1 and t[0]
        and any(is_date(str(c).strip()) for c in t[0] if c)
    ]
    header_only = [
        t for t in page_tables
        if t and len(t) == 1 and t[0]
        and not any(is_date(str(c).strip()) for c in t[0] if c)
        and any(" ".join(str(c).split()).lower() in HEADER_MAP for c in t[0] if c)
    ]
    all_txns, table_dates = [], set()
    if single_row_data and len(single_row_data) >= 3:
        merged = ([header_only[0][0]] if header_only else []) + [t[0] for t in single_row_data]
        if len(merged) >= 2:
            txns = parse_table(merged, bank, acct, fallback_year=yr)
            all_txns.extend(txns)
            table_dates = {t["date"] for t in txns}
        for t in parse_anz_cheque_text(all_text, fallback_year=yr):
            if t["date"] not in table_dates:
                all_txns.append(t)
    if not all_txns:
        for table in page_tables:
            if table and len(table) >= 2:
                all_txns.extend(parse_table(table, bank, acct, fallback_year=yr))
    if not all_txns:
        if bank == "anz" and acct == "credit_card":
            all_txns = parse_anz_credit_text(all_text, fallback_year=yr)
        elif bank == "westpac":
            all_txns = parse_westpac_rocket_text(all_text, fallback_year=yr)
            if not all_txns:
                all_txns = parse_anz_cheque_text(all_text, fallback_year=yr)
        elif bank == "commbank":
            all_txns = parse_cba_text(all_text, fallback_year=yr)
        elif bank == "hsbc" and acct == "credit_card":
            all_txns = parse_hsbc_credit_text(all_text, fallback_year=yr)
        elif bank == "anz":
            all_txns = parse_anz_cheque_text(all_text, fallback_year=yr)
        else:
            all_txns = parse_generic_text(all_text, acct)
    seen, unique = set(), []
    for t in all_txns:
        t["description"] = clean_desc(t.get("description", ""))
        key = (t["date"], t["description"], t["debit"], t["credit"])
        if key not in seen:
            seen.add(key)
            unique.append(t)
    return [{"meta": meta, "transactions": unique}]


# ============================================================
# -- SECTION B: INVOICE / RECEIPT LOGIC --
# ============================================================

def is_image(p: Path) -> bool:
    return p.suffix.lower() in IMG_EXTS


def is_pdf(p: Path) -> bool:
    return p.suffix.lower() in PDF_EXTS


def convert_image_to_pdf(image_path: Path, out_pdf: Path):
    img = Image.open(image_path).convert("RGB")
    img.save(str(out_pdf))


def preprocess_for_ocr(pil_img):
    if not _CV2_AVAILABLE:
        return pil_img
    img = cv2.cvtColor(np.array(pil_img.convert("RGB")), cv2.COLOR_RGB2BGR)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.bilateralFilter(gray, 9, 75, 75)
    thr = cv2.adaptiveThreshold(
        gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 31, 10
    )
    return Image.fromarray(thr)


def pdf_has_extractable_text(pdf_path: str) -> bool:
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page in pdf.pages[:2]:
                text = page.extract_text() or ""
                if len(text.strip()) > 50:
                    return True
        return False
    except Exception:
        return False


def extract_text_from_digital_pdf(pdf_path: str) -> str:
    chunks = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            chunks.append(page.extract_text() or "")
    return "\n".join(chunks)


def ocr_page_to_lines(pil_img, tesseract_cmd: str = None):
    if not _TESSERACT_AVAILABLE:
        raise RuntimeError(f"pytesseract could not be imported: {_TESSERACT_ERROR}")
    if tesseract_cmd:
        pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
    proc = preprocess_for_ocr(pil_img)
    data = pytesseract.image_to_data(proc, lang="eng", output_type=pytesseract.Output.DATAFRAME)
    data = data.dropna(subset=["text"])
    data["text"] = data["text"].astype(str).str.strip()
    data = data[(data["text"] != "") & (data["conf"] > 30)]
    if data.empty:
        return "", []
    lines = []
    for (_, _, _), grp in data.groupby(["block_num", "par_num", "line_num"], sort=True):
        grp = grp.sort_values("left")
        line = " ".join(grp["text"].tolist()).strip()
        if line:
            lines.append(line)
    return "\n".join(lines), lines


def extract_text_and_lines_from_scanned_pdf(pdf_path: str, poppler_bin: str = None,
                                             tesseract_cmd: str = None):
    if not _PDF2IMAGE_AVAILABLE:
        raise RuntimeError(f"pdf2image could not be imported: {_PDF2IMAGE_ERROR}")
    kwargs = {"dpi": 300}
    if poppler_bin:
        kwargs["poppler_path"] = poppler_bin
    pages = convert_from_path(str(pdf_path), **kwargs)
    all_lines, all_text_chunks = [], []
    for p in pages:
        page_text, page_lines = ocr_page_to_lines(p, tesseract_cmd=tesseract_cmd)
        all_text_chunks.append(page_text)
        all_lines.extend(page_lines)
    return "\n".join(all_text_chunks), all_lines, len(pages)


def extract_text_and_lines(pdf_path: str, poppler_bin: str = None, tesseract_cmd: str = None):
    if pdf_has_extractable_text(pdf_path):
        text = extract_text_from_digital_pdf(pdf_path)
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        return text, lines, "digital", 1
    text, lines, pages_count = extract_text_and_lines_from_scanned_pdf(
        pdf_path, poppler_bin=poppler_bin, tesseract_cmd=tesseract_cmd
    )
    return text, lines, "scanned", pages_count


# -- Invoice / Receipt field extractors ---------------------------------------

_DATE_FMTS = [
    r"\b(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{4})\b",
    r"\b(\d{4}[/\-\.]\d{1,2}[/\-\.]\d{1,2})\b",
    r"\b(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{4})\b",
    r"\b((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{1,2},?\s+\d{4})\b",
    r"\b(\d{1,2}[/\-]\d{1,2}[/\-]\d{2})\b",
]

_AMT_RE = r"\$?\s*([\d,]+(?:\.\d{2})?)"

_SKIP_SUMMARY = re.compile(
    r"^\s*(?:sub[\s\-]?total|grand\s+total|"
    r"total\s+(?:amount\s+)?(?:due|inc(?:l)?\.?\s*gst?|excl\.?\s*gst?|payable)?|"
    r"amount\s+(?:due|payable|owing)|"
    r"gst(?:\s+amount)?|tax(?:\s+amount)?|"
    r"balance\s+(?:due|owing|forward)|"
    r"discount|freight|shipping|postage|rounding|"
    r"deposit\s+paid|amount\s+paid|payment\s+received)\b",
    re.IGNORECASE,
)
_SKIP_META = re.compile(
    r"^\s*(?:invoice\s*(?:no|number|#|date)|bill\s+to|ship\s+to|sold\s+to|"
    r"customer|client|abn|acn|bsb|account\s+(?:no|name|number)|"
    r"payment\s+terms?|due\s+date|purchase\s+order|po\s+(?:no|number)|"
    r"your\s+ref|our\s+ref|page\s+\d|"
    r"^(?:qty|quantity|unit\s*price|description|item|code|product)$)\b",
    re.IGNORECASE,
)


def _norm_amt(raw: str) -> str:
    s = re.sub(r"[^\d.]", "", str(raw))
    try:
        return f"${float(s):,.2f}"
    except ValueError:
        return ""


def _first_date(text: str) -> str:
    for pat in _DATE_FMTS:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            return normalise_date(m.group(1).strip())
    return ""


def _after_label(line: str, label_re: re.Pattern) -> str:
    m = label_re.search(line)
    if m:
        return line[m.end():].strip(" :\t-")
    return ""


def _label_then_value(lines: List[str], label_re: re.Pattern,
                      value_fn=None, window: int = 40) -> str:
    if value_fn is None:
        value_fn = lambda t: t.strip()
    for i, line in enumerate(lines[:window]):
        if not label_re.search(line):
            continue
        remainder = _after_label(line, label_re)
        val = value_fn(remainder)
        if val:
            return val
        for nxt in lines[i + 1: i + 3]:
            nxt = nxt.strip()
            if nxt:
                val = value_fn(nxt)
                if val:
                    return val
    return ""


_INV_NO_RE = re.compile(
    r"(?:tax\s+)?invoice\s*(?:no|number|num|#)|"
    r"\binv\s*(?:no|#|num\b)|"
    r"receipt\s*(?:no|number|#)|"
    r"(?:reference|ref)\s*(?:no|number|#)?|"
    r"order\s*(?:no|number|#)",
    re.IGNORECASE,
)
_CODE_RE = re.compile(
    r"(?<!\w)([A-Z]{0,5}\d{3,}[\-\/]?[A-Z0-9]*|"
    r"[A-Z]{2,6}[\-\/]\d{2,}|"
    r"\d{4,}[\-][A-Z0-9\-]+)(?!\w)",
    re.IGNORECASE,
)


def _clean_inv_no(raw: str) -> str:
    tok = re.split(r"\s{2,}|\t", raw.strip())[0]
    tok = re.sub(r"[^A-Za-z0-9\-\/].*$", "", tok)
    if re.search(r"[A-Z0-9]", tok, re.IGNORECASE) and len(tok) >= 2:
        return tok.upper()
    return ""


def find_invoice_number(text: str, lines: List[str]) -> str:
    val = _label_then_value(lines, _INV_NO_RE, _clean_inv_no, window=50)
    if val:
        return val
    for line in lines[:25]:
        if re.search(r"^\s*\$|(?:total|balance|date|abn|bsb)", line, re.IGNORECASE):
            continue
        m = _CODE_RE.search(line)
        if m:
            return m.group(1).upper()
    return ""


_INV_DATE_RE = re.compile(
    r"invoice\s*date|date\s*(?:of\s*invoice|issued)?|"
    r"tax\s*date|issue\s*date|billing\s*date|document\s*date",
    re.IGNORECASE,
)


def find_invoice_date(text: str, lines: List[str]) -> str:
    val = _label_then_value(lines, _INV_DATE_RE, _first_date, window=50)
    if val:
        return val
    for line in lines[:15]:
        d = _first_date(line)
        if d:
            return d
    return _first_date(text)


_FROM_RE = re.compile(
    r"^(?:from|supplier|vendor|issued\s+by|billed\s+by)\s*[:\-]?\s*",
    re.IGNORECASE,
)
_COMPANY_SKIP = re.compile(
    r"^\s*(?:invoice|tax\s+invoice|receipt|abn|acn|gst|total|"
    r"bill\s+to|ship\s+to|sold\s+to|page\s+\d|"
    r"www\.|http|phone|ph\b|fax|email|"
    r"po\s+box|locked\s+bag|"
    r"(?:level|suite|shop|unit)\s+\d|"
    r"\d{1,4}\s+\w+\s+(?:st|rd|ave|dr|ln|blvd|street|road|avenue|drive|"
    r"lane|boulevard|way|place|court|crescent)\b)",
    re.IGNORECASE,
)
_DIGITS_ONLY = re.compile(r"^[\d\W\s]+$")


def find_company(lines: List[str], text: str = "") -> str:
    for line in lines[:30]:
        if _FROM_RE.match(line.strip()):
            val = _FROM_RE.sub("", line.strip()).strip()
            if val and not _DIGITS_ONLY.match(val):
                return val
    for ln in lines[:20]:
        ln = ln.strip()
        if len(ln) < 3:
            continue
        if _COMPANY_SKIP.match(ln) or _DIGITS_ONLY.match(ln):
            continue
        if re.match(r"^[\d/\-\.]+$", ln):
            continue
        return ln
    for i, ln in enumerate(lines[:35]):
        if re.search(r"\babn\b", ln, re.IGNORECASE) and i > 0:
            candidate = lines[i - 1].strip()
            if candidate and not _DIGITS_ONLY.match(candidate):
                return candidate
    return ""


def find_abn(text: str) -> str:
    m = re.search(
        r"\babn\s*[:\-]?\s*(\d{2}\s?\d{3}\s?\d{3}\s?\d{3})\b",
        text, re.IGNORECASE,
    )
    return normalise_abn(m.group(1).strip()) if m else ""


_GST_PATS = [
    r"gst\s+(?:amount|charged|included|component)\s*[:\-]?\s*" + _AMT_RE,
    r"gst\s*[:\-]\s*" + _AMT_RE,
    r"tax\s+(?:amount|charged)\s*[:\-]?\s*" + _AMT_RE,
    r"(?:inc(?:l)?\.?\s+gst|gst\s+inc(?:l)?\.?)\s*[:\-]?\s*" + _AMT_RE,
    r"\bgst\b[^\n]{0,30}" + _AMT_RE,
]


def find_gst(text: str) -> str:
    for pat in _GST_PATS:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            return normalise_amount(m.group(1))
    return ""


_TOTAL_PATS = [
    r"(?:grand\s+)?total\s+(?:amount\s+)?(?:due|payable|inc(?:l)?\.?\s*gst?|including\s+gst)\s*[:\-]?\s*" + _AMT_RE,
    r"amount\s+(?:due|payable|owing)\s*[:\-]?\s*" + _AMT_RE,
    r"balance\s+(?:due|owing|payable)\s*[:\-]?\s*" + _AMT_RE,
    r"invoice\s+total\s*[:\-]?\s*" + _AMT_RE,
    r"total\s+(?:inc(?:l)?\.?\s*gst?)?\s*[:\-]?\s*" + _AMT_RE,
    r"(?:grand\s+)?total\s*[:\-]?\s*" + _AMT_RE,
]


def find_total_amount(text: str) -> str:
    for pat in _TOTAL_PATS:
        hits = list(re.finditer(pat, text, re.IGNORECASE))
        if hits:
            return normalise_amount(hits[-1].group(1))
    return ""


def extract_line_items(lines: List[str], text: str = "") -> List[dict]:
    items: List[dict] = []
    _hdr_re = re.compile(
        r"(?P<desc>desc(?:ription)?|item|product|particulars|service)"
        r".*?(?P<qty>qty|quantity|units?)?"
        r".*?(?P<unit>unit\s*(?:price|cost|rate)|rate|price\s*each)?"
        r".*?(?P<amt>amount|total|price|charge)",
        re.IGNORECASE,
    )
    _money_hit = re.compile(r"\$?\s*([\d,]+\.\d{2})")

    def _parse_qty_unit(desc_raw, money_list, line_total):
        qty, unit_price = "", ""
        desc = desc_raw.strip()
        qm = re.search(r"(?:^|\s)qty\s*:?\s*(\d+(?:\.\d+)?)(?=\s|$)", desc, re.IGNORECASE)
        if qm:
            qty = qm.group(1)
            desc = (desc[: qm.start()] + desc[qm.end():]).strip()
        if not qty:
            qm = re.match(r"^(\d+(?:\.\d+)?)\s+(.+)", desc)
            if qm and not re.match(r"^\d{4}[-/]", qm.group(1)):
                qty = qm.group(1)
                desc = qm.group(2).strip()
        if not qty:
            qm = re.search(r"\s+[x-]\s*(\d+(-:\.\d+)-)\s*$", desc, re.IGNORECASE)
            if qm:
                qty = qm.group(1)
                desc = desc[: qm.start()].strip()
        if len(money_list) >= 2:
            candidate = money_list[-2]
            if qty:
                try:
                    calc = float(qty) * float(candidate.replace(",", ""))
                    total_f = float(line_total.replace(",", ""))
                    if total_f > 0 and abs(calc - total_f) / total_f <= 0.02:
                        unit_price = candidate
                except (ValueError, ZeroDivisionError):
                    pass
            else:
                try:
                    up_f = float(candidate.replace(",", ""))
                    tot_f = float(line_total.replace(",", ""))
                    if 0 < up_f <= tot_f:
                        unit_price = candidate
                except ValueError:
                    pass
        return desc, qty, unit_price

    header_idx = None
    for i, line in enumerate(lines):
        if _hdr_re.search(line) and i < len(lines) - 2:
            header_idx = i
            break
    if header_idx is not None:
        for line in lines[header_idx + 1:]:
            ls = line.rstrip()
            if not ls:
                continue
            if _SKIP_SUMMARY.match(ls):
                break
            if _SKIP_META.match(ls):
                continue
            money = _money_hit.findall(ls)
            if not money:
                continue
            line_total = money[-1]
            first_money_pos = _money_hit.search(ls).start()
            desc_raw = ls[:first_money_pos].strip()
            if len(desc_raw) < 2:
                continue
            desc, qty, unit_price = _parse_qty_unit(desc_raw, money, line_total)
            if len(desc) < 2:
                continue
            items.append({
                "Item description": desc, "Qty": qty,
                "Unit Price": _norm_amt(unit_price) if unit_price else "",
                "Amount": _norm_amt(line_total),
            })

    if not items:
        _item_re = re.compile(
            r"^(.{3,70}?)\s+"
            r"(?:(\d+(?:\.\d+)?)\s+\$?([\d,]+\.\d{2})\s+)?"
            r"\$?([\d,]+\.\d{2})\s*$",
            re.IGNORECASE,
        )
        for line in lines:
            ls = line.strip()
            if not ls or _SKIP_SUMMARY.match(ls) or _SKIP_META.match(ls):
                continue
            m = _item_re.match(ls)
            if not m:
                continue
            money_on_line = _money_hit.findall(ls)
            if not money_on_line:
                continue
            total_raw = money_on_line[-1]
            first_pos = _money_hit.search(ls).start()
            desc_raw = ls[:first_pos].strip()
            if not desc_raw:
                desc_raw = (m.group(1) or "").strip()
            if re.search(r"(?:invoice|date|abn|bsb|page)", desc_raw, re.IGNORECASE):
                continue
            desc, qty, unit_price = _parse_qty_unit(desc_raw, money_on_line, total_raw)
            if len(desc) < 2:
                continue
            items.append({
                "Item description": desc, "Qty": qty,
                "Unit Price": _norm_amt(unit_price) if unit_price else "",
                "Amount": _norm_amt(total_raw),
            })

    if not items:
        _loose = re.compile(r"^(.{2,70}?)\s+\$?([\d,]+\.\d{2})\s*$")
        for line in lines:
            ls = line.strip()
            if not ls or _SKIP_SUMMARY.match(ls) or _SKIP_META.match(ls):
                continue
            m = _loose.match(ls)
            if not m:
                continue
            desc_raw = m.group(1).strip()
            if re.search(r"(?:invoice|date|abn|bsb)", desc_raw, re.IGNORECASE):
                continue
            total_raw = m.group(2)
            money_on_line = _money_hit.findall(ls)
            desc, qty, unit_price = _parse_qty_unit(desc_raw, money_on_line, total_raw)
            items.append({
                "Item description": desc, "Qty": qty,
                "Unit Price": _norm_amt(unit_price) if unit_price else "",
                "Amount": _norm_amt(total_raw),
            })

    seen: set = set()
    unique: List[dict] = []
    for it in items:
        key = (it["Item description"], it["Amount"])
        if key not in seen:
            seen.add(key)
            unique.append(it)
    return unique or [{"Item description": "", "Qty": "", "Unit Price": "", "Amount": ""}]


def extract_fields(text: str, lines: List[str]) -> dict:
    return {
        "Invoice number": clean_invoice_number(find_invoice_number(text, lines)),
        "Invoice date":   normalise_date(find_invoice_date(text, lines)),
        "Company":        clean_company(find_company(lines, text)),
        "ABN":            find_abn(text),
        "Gst":            normalise_gst(find_gst(text)),
        "Invoice total":  normalise_amount(find_total_amount(text)),
    }


def classify_document_type(text: str, source_is_image: bool) -> str:
    if source_is_image:
        return "receipt"
    if is_bank_statement(text):
        return "bank_statement"
    return "invoice"


# -- Excel writers -------------------------------------------------------------

def _bank_excel_bytes(bank_results: list) -> bytes:
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is not installed.")

    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    SUB_HDR_FILL = PatternFill("solid", start_color="2E75B6")
    DAT_FONT   = Font(size=9, name="Arial")
    DEB_FILL   = PatternFill("solid", start_color="FFF0F0")
    CRE_FILL   = PatternFill("solid", start_color="F0FFF0")
    ALT_FILL   = PatternFill("solid", start_color="F5F9FF")
    THIN       = Side(style="thin", color="CCCCCC")
    BDR        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    AMT_FMT    = "#,##0.00"

    COLS = [
        ("Bank", 18), ("Account Type", 18), ("Date", 13),
        ("Description", 52), ("Reference", 16),
        ("Debit (AUD)", 14), ("Credit (AUD)", 14), ("Balance (AUD)", 15),
        ("Cumul. Debit", 14), ("Cumul. Credit", 14),
        ("Period", 26), ("Source File", 36),
    ]
    n_cols = len(COLS)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Transactions"

    for ci, (hdr, w) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.font = HDR_FONT
        c.fill = SUB_HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BDR
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 20

    current_row = 2
    for result, source_filename in bank_results:
        meta = result["meta"]
        txns = result["transactions"]
        bank_lbl = meta["bank"].upper() if meta["bank"] != "unknown" else "UNKNOWN BANK"
        acct_lbl = (meta["account_type"].replace("_", " ").title()
                    if meta["account_type"] != "unknown" else "Unknown Account")
        period = ""
        if meta.get("period_start"):
            period = meta["period_start"]
            if meta.get("period_end"):
                period += f"  -  {meta['period_end']}"

        for i, txn in enumerate(txns):
            r = current_row
            if txn.get("debit"):
                fill = DEB_FILL
            elif txn.get("credit"):
                fill = CRE_FILL
            elif i % 2 == 0:
                fill = ALT_FILL
            else:
                fill = None

            def wc(col, val, fmt=None, align="left", _r=r, _fill=fill):
                c = ws.cell(row=_r, column=col, value=val)
                c.font = DAT_FONT
                c.border = BDR
                c.alignment = Alignment(horizontal=align)
                if _fill:
                    c.fill = _fill
                if fmt:
                    c.number_format = fmt
                return c

            wc(1, bank_lbl, align="center")
            wc(2, acct_lbl, align="center")
            wc(3, txn.get("date"), align="center")
            wc(4, txn.get("description"))
            wc(5, txn.get("reference"), align="center")
            wc(6, txn.get("debit"), fmt=AMT_FMT, align="right")
            wc(7, txn.get("credit"), fmt=AMT_FMT, align="right")
            wc(8, txn.get("balance"), fmt=AMT_FMT, align="right")
            prev = r - 1
            for col, src_col in ((9, "F"), (10, "G")):
                if r == 2:
                    formula = f"=IFERROR({src_col}{r},0)"
                else:
                    cum_col = get_column_letter(col)
                    formula = f"=IFERROR({cum_col}{prev}+{src_col}{r},{cum_col}{prev})"
                c = ws.cell(row=r, column=col, value=formula)
                c.number_format = AMT_FMT
                c.font = DAT_FONT
                c.border = BDR
                c.alignment = Alignment(horizontal="right")
                if fill:
                    c.fill = fill
            wc(11, period, align="center")
            wc(12, source_filename)
            ws.row_dimensions[r].height = 14
            current_row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{current_row - 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _inv_rec_excel_bytes(df: pd.DataFrame) -> bytes:
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is not installed.")

    THIN_S     = Side(style="thin", color="D0D0D0")
    BDR_S      = Border(left=THIN_S, right=THIN_S, top=THIN_S, bottom=THIN_S)
    HDR_FILL   = PatternFill("solid", start_color="1A3A5C")
    HDR_FONT_L = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    INV_FILL   = PatternFill("solid", start_color="EAF3FB")
    REC_FILL   = PatternFill("solid", start_color="E8F8F0")
    ALT_INV    = PatternFill("solid", start_color="D6E8F5")
    ALT_REC    = PatternFill("solid", start_color="D0F0E0")
    DAT_FONT_L = Font(size=9, name="Arial")
    TTL_FONT_L = Font(bold=True, size=13, name="Arial", color="1A3A5C")
    SUB_FONT_L = Font(size=9, name="Arial", italic=True, color="666666")

    COLS = list(df.columns)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices & Receipts"

    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    ws["A1"] = "Invoices & Receipts - Extraction Report"
    ws["A1"].font = TTL_FONT_L
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{get_column_letter(len(COLS))}2")
    ws["A2"] = (
        f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}   |   "
        f"Total rows: {len(df)}   |   "
        f"Invoices: {(df['Document Type'] == 'Invoice').sum()}   |   "
        f"Receipts: {(df['Document Type'] == 'Receipt').sum()}"
    )
    ws["A2"].font = SUB_FONT_L
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    HDR_ROW = 4
    CENTER_COLS = {"Document Type", "Invoice date", "Mode", "Pages", "ABN", "Qty", "Invoice number"}
    RIGHT_COLS  = {"Amount", "GST", "Invoice total", "Unit Price"}
    COL_WIDTHS  = {
        "Document Type": 13, "Invoice number": 16, "Invoice date": 13,
        "Company": 30, "ABN": 16,
        "Item description": 45, "Qty": 8, "Unit Price": 13,
        "Amount": 13, "GST": 11, "Invoice total": 14,
        "Mode": 9, "Pages": 7, "Source file": 36,
    }

    for ci, col_name in enumerate(COLS, 1):
        c = ws.cell(HDR_ROW, ci, col_name)
        c.font = HDR_FONT_L
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BDR_S
    ws.row_dimensions[HDR_ROW].height = 22

    for ri, (_, row) in enumerate(df.iterrows()):
        r = HDR_ROW + 1 + ri
        is_inv = str(row.get("Document Type", "")).strip().lower() == "invoice"
        fill = (INV_FILL if is_inv else REC_FILL) if ri % 2 == 0 else (ALT_INV if is_inv else ALT_REC)
        for ci, col_name in enumerate(COLS, 1):
            val = row.get(col_name, "")
            cell = ws.cell(r, ci, val if val != "" else None)
            cell.font = DAT_FONT_L
            cell.fill = fill
            cell.border = BDR_S
            if col_name in CENTER_COLS:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name in RIGHT_COLS:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 15

    for ci, col_name in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col_name, 14)

    ws.freeze_panes = f"A{HDR_ROW + 1}"
    ws.auto_filter.ref = f"A{HDR_ROW}:{get_column_letter(len(COLS))}{HDR_ROW + len(df)}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# -- PUBLIC API --
# ============================================================

def get_dependency_status() -> dict:
    """Return availability and any import errors for optional dependencies."""
    return {
        "cv2":       {"available": _CV2_AVAILABLE,       "error": _CV2_ERROR},
        "pdf2image": {"available": _PDF2IMAGE_AVAILABLE, "error": _PDF2IMAGE_ERROR},
        "pytesseract": {"available": _TESSERACT_AVAILABLE, "error": _TESSERACT_ERROR},
        "openpyxl":  {"available": _OPENPYXL_AVAILABLE,  "error": _OPENPYXL_ERROR},
    }


def process_files(
    uploaded_files: List[Tuple[str, bytes]],
    tesseract_cmd: str = None,
    poppler_bin: str = None,
    progress_callback=None,
) -> dict:
    """
    Process a list of (filename, file_bytes) tuples.

    Returns:
        {
          "inv_rec_df":      pd.DataFrame | None,
          "bank_results":    list of (result_dict, filename),
          "bank_excel":      bytes | None,
          "inv_rec_excel":   bytes | None,
          "inv_rec_csv":     str | None,
          "summary": {
              "total": int, "bank_txns": int,
              "invoices": int, "receipts": int,
              "errors": int, "ocr_used": int,
              "log": list[str],
          },
        }
    """
    log: List[str] = []
    inv_rec_rows = []
    bank_results = []
    bank_count = inv_count = rec_count = error_count = ocr_used = 0

    tmpdir = Path(tempfile.mkdtemp())
    try:
        image_originated: set = set()
        pdf_paths: List[Path] = []

        # Write uploaded files to temp dir; convert images - PDF
        for filename, file_bytes in uploaded_files:
            p = tmpdir / filename
            p.write_bytes(file_bytes)
            if is_image(p):
                pdf_path = tmpdir / f"{p.stem}.pdf"
                convert_image_to_pdf(p, pdf_path)
                pdf_paths.append(pdf_path)
                image_originated.add(pdf_path.name)
            elif is_pdf(p):
                pdf_paths.append(p)

        total = len(pdf_paths)
        for idx, pdf in enumerate(pdf_paths):
            source_is_image = pdf.name in image_originated
            if progress_callback:
                progress_callback(idx, total, pdf.name)
            log.append(f"Processing: {pdf.name}")
            try:
                text, lines, mode, pages_count = extract_text_and_lines(
                    str(pdf), poppler_bin=poppler_bin, tesseract_cmd=tesseract_cmd
                )
                if mode == "scanned":
                    ocr_used += 1
                doc_type = classify_document_type(text, source_is_image)
                log.append(f"  - {doc_type.upper()} | mode={mode} | pages={pages_count}")

                if doc_type == "bank_statement":
                    results = extract_from_pdf_bank(str(pdf))
                    for result in results:
                        bank_results.append((result, pdf.name))
                        meta = result["meta"]
                        n_txns = len(result["transactions"])
                        bank_count += n_txns
                        log.append(f"  - Bank: {meta['bank']} | Type: {meta['account_type']} | Txns: {n_txns}")
                else:
                    label = "Invoice" if doc_type == "invoice" else "Receipt"
                    header = extract_fields(text, lines)
                    items = extract_line_items(lines, text)
                    for it in items:
                        inv_rec_rows.append({
                            "Document Type":    label,
                            "Invoice number":   header.get("Invoice number", ""),
                            "Invoice date":     header.get("Invoice date", ""),
                            "Company":          header.get("Company", ""),
                            "ABN":              header.get("ABN", ""),
                            "Item description": it.get("Item description", ""),
                            "Qty":              it.get("Qty", ""),
                            "Unit Price":       it.get("Unit Price", ""),
                            "Amount":           it.get("Amount", ""),
                            "GST":              header.get("Gst", ""),
                            "Invoice total":    header.get("Invoice total", ""),
                            "Mode":             mode,
                            "Pages":            pages_count,
                            "Source file":      pdf.name,
                        })
                    if doc_type == "invoice":
                        inv_count += 1
                    else:
                        rec_count += 1
                    log.append(f"  - {label} | items={len(items)}")

            except Exception as exc:
                error_count += 1
                log.append(f"  ERROR: {exc}")
                inv_rec_rows.append({
                    "Document Type": "Error", "Invoice number": "", "Invoice date": "",
                    "Company": "", "ABN": "",
                    "Item description": f"ERROR: {exc}", "Qty": "", "Unit Price": "",
                    "Amount": "", "GST": "", "Invoice total": "",
                    "Mode": "error", "Pages": "", "Source file": pdf.name,
                })

        # Build outputs
        inv_rec_df = None
        inv_rec_excel = None
        inv_rec_csv = None
        if inv_rec_rows:
            COLS_ORDER = [
                "Document Type", "Invoice number", "Invoice date",
                "Company", "ABN",
                "Item description", "Qty", "Unit Price",
                "Amount", "GST", "Invoice total",
                "Mode", "Pages", "Source file",
            ]
            df = pd.DataFrame(inv_rec_rows)
            for c in COLS_ORDER:
                if c not in df.columns:
                    df[c] = ""
            df = df[COLS_ORDER]
            df = _apply_column_validators(df)
            inv_rec_df = df
            if _OPENPYXL_AVAILABLE:
                inv_rec_excel = _inv_rec_excel_bytes(df)
            inv_rec_csv = df.to_csv(index=False)

        bank_excel = None
        if bank_results and _OPENPYXL_AVAILABLE:
            bank_excel = _bank_excel_bytes(bank_results)

        return {
            "inv_rec_df":    inv_rec_df,
            "bank_results":  bank_results,
            "bank_excel":    bank_excel,
            "inv_rec_excel": inv_rec_excel,
            "inv_rec_csv":   inv_rec_csv,
            "summary": {
                "total":     total,
                "bank_txns": bank_count,
                "invoices":  inv_count,
                "receipts":  rec_count,
                "errors":    error_count,
                "ocr_used":  ocr_used,
                "log":       log,
            },
        }
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)
