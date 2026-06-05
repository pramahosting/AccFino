"""
excel_exporter.py — HSLedger Trading Module
Builds the final consolidated Excel tax report.

Sheet layout:
  1. Summary          — FY-level tax summary across all brokers
  2. CGT Disposals    — All disposal rows (one per lot consumed)
  3. Income           — Dividends, interest, lending income
  4. Missing Buys     — Unmatched sells requiring user action (orange flag)
  5. Open Positions   — Remaining lots still held (unrealised)
  6. {Broker}         — Per-broker raw disposal detail (one sheet per broker)
  7. Duplicates       — Rows removed during dedup (audit trail)
"""

from __future__ import annotations

import io
import os
from datetime import date
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from equity.equity_engine import (
    DisposalRow, FYSummary, IncomeRow, MissingBuyFlag,
    disposals_to_df, income_to_df, missing_to_df, summary_to_df,
)

# ── Colour palette ────────────────────────────────────────────────────────────
C_NAVY     = "1F3864"
C_BLUE_LT  = "DBEAFE"
C_GREEN    = "D1FAE5"
C_ORANGE   = "FEF3C7"
C_RED_LT   = "FEE2E2"
C_GREY_LT  = "F3F4F6"
C_WHITE    = "FFFFFF"
C_BLACK    = "000000"
C_BORDER   = "CBD5E1"

HEADER_FONT  = Font(name="Arial", bold=True, color=C_WHITE, size=10)
BODY_FONT    = Font(name="Arial", size=10)
BOLD_FONT    = Font(name="Arial", bold=True, size=10)
TITLE_FONT   = Font(name="Arial", bold=True, size=13)
SUBHEAD_FONT = Font(name="Arial", bold=True, size=11, color=C_NAVY)

NAVY_FILL   = PatternFill("solid", fgColor=C_NAVY)
BLUE_FILL   = PatternFill("solid", fgColor=C_BLUE_LT)
GREEN_FILL  = PatternFill("solid", fgColor=C_GREEN)
ORANGE_FILL = PatternFill("solid", fgColor=C_ORANGE)
RED_FILL    = PatternFill("solid", fgColor=C_RED_LT)
GREY_FILL   = PatternFill("solid", fgColor=C_GREY_LT)
ALT_FILL    = PatternFill("solid", fgColor="EEF2F7")

def _thin_border(color: str = C_BORDER) -> Border:
    s = Side(style="thin", color=color)
    return Border(bottom=s)

def _all_border(color: str = C_BORDER) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


# ── Sheet formatting helpers ──────────────────────────────────────────────────

def _apply_header_row(ws, row_num: int = 1, fill=None) -> None:
    fill = fill or NAVY_FILL
    for cell in ws[row_num]:
        cell.font      = HEADER_FONT
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _thin_border()
    ws.row_dimensions[row_num].height = 28


def _apply_body_rows(ws, start_row: int = 2, alt_fill=True, row_fill_override=None) -> None:
    for i, row in enumerate(ws.iter_rows(min_row=start_row), start_row):
        use_fill = row_fill_override(i) if row_fill_override else (ALT_FILL if (i % 2 == 0 and alt_fill) else PatternFill())
        for cell in row:
            cell.font      = BODY_FONT
            cell.fill      = use_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = _thin_border()


def _auto_width(ws, min_w: int = 8, max_w: int = 44) -> None:
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(min_w, min(w + 3, max_w))


def _freeze_and_filter(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions


def _write_df_to_sheet(ws, df: pd.DataFrame, header_fill=None, row_fill_fn=None) -> None:
    """Write a DataFrame to a worksheet starting at row 1."""
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    _apply_header_row(ws, row_num=1, fill=header_fill)
    _apply_body_rows(ws, start_row=2, row_fill_override=row_fill_fn)
    _auto_width(ws)
    _freeze_and_filter(ws)


# ── Summary sheet ─────────────────────────────────────────────────────────────

def _write_summary_sheet(ws, fy_summaries: dict[str, FYSummary], report_date: str) -> None:
    ws.title = "Summary"

    # Title block
    ws["A1"] = "HSLedger — Equity CGT Tax Report"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated: {report_date}"
    ws["A2"].font = Font(name="Arial", size=10, color="6B7280")
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers row 4
    headers = [
        "Financial Year", "Disposals", "Total Proceeds ($)", "Total Cost Base ($)",
        "Gross Gains ($)", "Gross Losses ($)", "CGT Discount ($)",
        "Net Cap Gain (pre CF) ($)", "Carry-Forward In ($)",
        "Net Taxable Gain ($)", "Income Events", "Total Income ($)",
        "Missing Buys ⚠", "Brokers",
    ]
    ws.append([])   # row 3 blank
    ws.append(headers)
    _apply_header_row(ws, row_num=4)

    for s in sorted(fy_summaries.values(), key=lambda x: x.financial_year):
        row = [
            s.financial_year, s.disposal_count,
            s.total_proceeds, s.total_cost_base,
            s.gross_gains, s.gross_losses, s.cgt_discount_applied,
            s.net_capital_gain_pre, s.carry_forward_in,
            s.net_after_carryforward, s.income_count, s.total_income,
            len(s.missing_buys), ", ".join(s.brokers),
        ]
        ws.append(row)

    # Style data rows
    for i, row in enumerate(ws.iter_rows(min_row=5), 5):
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        # Flag rows with missing buys in orange
        miss_cell = row[12]
        if miss_cell.value and int(miss_cell.value) > 0:
            for cell in row:
                cell.fill = ORANGE_FILL
        else:
            for cell in row:
                cell.fill = fill
        for cell in row:
            cell.font      = BODY_FONT
            cell.alignment = Alignment(horizontal="right" if isinstance(cell.value, (int, float)) else "left", vertical="center")
            cell.border    = _thin_border()

    # Number formats
    money_cols = [3, 4, 5, 6, 7, 8, 9, 10, 12]   # 1-indexed
    for row in ws.iter_rows(min_row=5):
        for idx, cell in enumerate(row, 1):
            if idx in money_cols:
                cell.number_format = '#,##0.00'
            if idx in [2, 11, 13]:
                cell.number_format = '#,##0'

    _auto_width(ws)
    ws.freeze_panes = "A5"
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[4].height = 28


# ── Main export function ──────────────────────────────────────────────────────

def _write_resolution_progress_sheet(
    ws,
    resolution_log: list[dict],
    missing_flags: list[MissingBuyFlag],
) -> None:
    """Write the Missing Buy Resolution Progress sheet."""
    ws.title = "Resolution Progress"

    # Build unified view: start from missing_flags, enrich with log entries
    rows = []
    log_by_code_date: dict[tuple, list[dict]] = {}
    for entry in resolution_log:
        key = (entry.get("code", ""), str(entry.get("sell_date", "")))
        log_by_code_date.setdefault(key, []).append(entry)

    # Current missing flags (after last run)
    after_map: dict[tuple, float] = {
        (m.code, str(m.disposal_date)): m.qty_unmatched
        for m in missing_flags
    }

    # Collect all keys ever logged
    all_keys: set[tuple] = set(log_by_code_date.keys())
    for m in missing_flags:
        all_keys.add((m.code, str(m.disposal_date)))

    for key in sorted(all_keys):
        code, sell_date_str = key
        entries = log_by_code_date.get(key, [])
        latest  = entries[-1] if entries else {}

        before = float(latest.get("qty_unmatched_before", 0)) if latest else 0.0
        added  = float(latest.get("qty_added", 0)) if latest else 0.0
        after  = float(latest.get("qty_unmatched_after", after_map.get(key, before)))

        if after == 0.0:
            status = "Resolved"
        elif after < before:
            status = "Partially Resolved"
        else:
            status = "Unresolved"

        lot_ids = "; ".join(latest.get("lot_ids_added", [])) if latest else ""

        rows.append({
            "Code":                 code,
            "Sell Date":            sell_date_str,
            "Qty Unmatched Before": before,
            "Qty Added":            added,
            "Qty Unmatched After":  after,
            "Status":               status,
            "Lot IDs Added":        lot_ids,
            "Last Updated":         latest.get("timestamp", ""),
        })

    if not rows:
        ws.append(["No missing buy resolution activity recorded yet."])
        ws["A1"].font = Font(name="Arial", italic=True, color="6B7280")
        return

    df = pd.DataFrame(rows)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    _apply_header_row(ws, row_num=1, fill=PatternFill("solid", fgColor="7C3AED"))

    def _res_fill(row_idx: int):
        rows_list = list(ws.iter_rows(min_row=row_idx, max_row=row_idx))
        if not rows_list:
            return PatternFill()
        status_cell = rows_list[0][5]   # "Status" column (index 5)
        val = str(status_cell.value or "")
        if val == "Resolved":
            return GREEN_FILL
        if val == "Partially Resolved":
            return ORANGE_FILL
        if val == "Unresolved":
            return RED_FILL
        return ALT_FILL if row_idx % 2 == 0 else PatternFill()

    _apply_body_rows(ws, start_row=2, row_fill_override=_res_fill)
    _auto_width(ws)
    _freeze_and_filter(ws)


def export_to_excel(
    disposals:      list[DisposalRow],
    income_events:  list[IncomeRow],
    fy_summaries:   dict[str, FYSummary],
    missing_flags:  list[MissingBuyFlag],
    open_positions: dict[str, Any] | None = None,
    duplicates_df:  pd.DataFrame | None = None,
    resolution_log: list[dict] | None = None,
    output_path:    str | None = None,
) -> bytes:
    """
    Build full Excel report and return as bytes (also writes to output_path if given).

    Parameters
    ----------
    disposals       : CGT disposal rows from equity_engine
    income_events   : Income rows
    fy_summaries    : FY summary dict
    missing_flags   : Unmatched sell flags
    open_positions  : dict of code → deque[Lot] (remaining holdings)
    duplicates_df   : DataFrame of removed duplicates (from deduplicator)
    output_path     : Optional file path to save .xlsx

    Returns
    -------
    bytes — Excel file content
    """
    import datetime
    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet
    report_date = datetime.date.today().strftime("%d/%m/%Y")

    # ── 1. Summary ────────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    _write_summary_sheet(ws_sum, fy_summaries, report_date)

    # ── 2. CGT Disposals ─────────────────────────────────────────────────────
    ws_disp = wb.create_sheet("CGT Disposals")
    disp_df = disposals_to_df(disposals)
    if not disp_df.empty:
        def disp_fill(row_idx: int):
            row_data = list(ws_disp.iter_rows(min_row=row_idx, max_row=row_idx))[0]
            # Green if discount eligible, normal alt otherwise
            # "Discount Eligible" is column index 14 (0-based) after the "Account" column was added
            disc_val = row_data[14].value   # "Discount Eligible" column index
            if disc_val is True:
                return GREEN_FILL
            return ALT_FILL if row_idx % 2 == 0 else PatternFill()
        _write_df_to_sheet(ws_disp, disp_df, row_fill_fn=disp_fill)
    else:
        ws_disp.append(["No disposal events in this period."])

    # ── 3. Income ─────────────────────────────────────────────────────────────
    ws_inc = wb.create_sheet("Income")
    inc_df = income_to_df(income_events)
    if not inc_df.empty:
        _write_df_to_sheet(ws_inc, inc_df, header_fill=PatternFill("solid", fgColor="065F46"))
    else:
        ws_inc.append(["No income events in this period."])

    # ── 4. Missing Buys ───────────────────────────────────────────────────────
    ws_miss = wb.create_sheet("⚠ Missing Buys")
    miss_df = missing_to_df(missing_flags)
    if not miss_df.empty:
        _write_df_to_sheet(ws_miss, miss_df, header_fill=PatternFill("solid", fgColor="92400E"),
                           row_fill_fn=lambda _: ORANGE_FILL)
        # Add instruction cell
        last_row = ws_miss.max_row + 2
        ws_miss.cell(last_row, 1).value = (
            "ACTION REQUIRED: Add missing purchase details to cost_base_history.csv "
            "or use the manual entry prompt, then re-run the module."
        )
        ws_miss.cell(last_row, 1).font = Font(name="Arial", bold=True, color="92400E", size=10)
        ws_miss.merge_cells(f"A{last_row}:G{last_row}")
    else:
        ws_miss.append(["✓ No missing buy transactions detected."])
        ws_miss["A1"].font = Font(name="Arial", bold=True, color="065F46")

    # ── 4b. Missing Buy Resolution Progress ──────────────────────────────────
    ws_res = wb.create_sheet("Resolution Progress")
    _write_resolution_progress_sheet(ws_res, resolution_log or [], missing_flags)

    # ── 5. Open Positions ─────────────────────────────────────────────────────
    ws_open = wb.create_sheet("Open Positions")
    if open_positions:
        rows_op = []
        for key, queue in open_positions.items():
            # Keys are "account_id::code" when multiple accounts are present
            if "::" in key:
                account_label, code = key.split("::", 1)
            else:
                account_label, code = "—", key
            for lot in queue:
                rows_op.append({
                    "Account":          account_label or "—",
                    "Asset Code":       code,
                    "Qty Held":         round(lot.qty, 4),
                    "Cost/Unit ($)":    round(lot.cost_per_unit, 4),
                    "Total Cost ($)":   round(lot.qty * lot.cost_per_unit, 2),
                    "Acquisition Date": lot.trade_date,
                    "Settlement Date":  lot.settlement_date,
                    "Days Held":        (date.today() - lot.trade_date).days,
                    "Discount Eligible": (date.today() - lot.trade_date).days > 365,
                    "Broker":           lot.broker,
                })
        op_df = pd.DataFrame(rows_op)
        if not op_df.empty:
            _write_df_to_sheet(ws_open, op_df, header_fill=PatternFill("solid", fgColor="1E40AF"))
        else:
            ws_open.append(["No open positions."])
    else:
        ws_open.append(["No open positions data provided."])

    # ── 6. Per-broker sheets ──────────────────────────────────────────────────
    if not disp_df.empty and "Broker" in disp_df.columns:
        for broker in sorted(disp_df["Broker"].dropna().unique()):
            safe_name = str(broker)[:28].replace("/", "-")
            ws_b = wb.create_sheet(safe_name)
            broker_df = disp_df[disp_df["Broker"] == broker].reset_index(drop=True)
            _write_df_to_sheet(ws_b, broker_df, header_fill=PatternFill("solid", fgColor="3730A3"))

    # ── 7. Duplicates (audit trail) ───────────────────────────────────────────
    if duplicates_df is not None and not duplicates_df.empty:
        ws_dup = wb.create_sheet("Duplicates (Audit)")
        _write_df_to_sheet(ws_dup, duplicates_df, header_fill=PatternFill("solid", fgColor="6B7280"))

    # ── Save ──────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    if output_path:
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        with open(output_path, "wb") as f:
            f.write(content)
        print(f"[exporter] Report saved -> {output_path}")

    return content
